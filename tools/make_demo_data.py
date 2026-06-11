# -*- coding: utf-8 -*-
"""產生展示用範例資料（全虛構人物/廠商/事件，僅保留與真實檔案相同的「格式結構」）"""
import os
import random
from openpyxl import Workbook

random.seed(42)
OUT = os.path.join(os.path.dirname(__file__), "..", "sample-data")
os.makedirs(OUT, exist_ok=True)


def phone():
    return f"09{random.randint(0,9)}{random.randint(0,9)}-{random.randint(100000,999999)}"


def ext():
    return str(random.randint(71001, 71999))

# ---------- 1. 通訊錄 ----------

wb = Workbook()
ws = wb.active
ws.title = "通訊錄20260301"

STAFF_L = [("處本部", ["王志遠", "李承恩"]),
           ("系統服務科(8人)", ["陳雅筑", "林冠宇", "張書豪", "黃詩涵", "吳承翰", "蔡佩珊", "許育誠", "鄭如萱"])]
STAFF_R = [("", []),
           ("基礎設施科(6人)", ["劉建宏", "楊曉彤", "周治平", "蘇芳儀", "高瑞陽", "簡millie"])]
STAFF_R[1] = ("基礎設施科(6人)", ["劉建宏", "楊曉彤", "周治平", "蘇芳儀", "高瑞陽", "簡郁芬"])

VENDORS_L = [("晨曦資訊\n(機房OP值班24HR)", ["范植偉", "郭哲安", "值班手機"]),
             ("北辰軟體\n(人事系統)", ["馮思婷", "趙國銘"]),
             ("星辰科技\n(多媒體看板)", ["錢柏均", "孫嘉玲"]),
             ("雲嶺電信\n(官網)", ["洪建文", "曾若彤"])]
VENDORS_R = [("宏宇系統\n(虛擬機)", ["宏宇值班"]),
             ("鼎峰工程\n(門禁系統)", ["呂偉誠"]),
             ("嵐光科技\n(監視系統)", ["江雨潔"]),
             ("藍鯨網路\n(網路)", ["藍鯨值班", "石明峰"]),
             ("青嶼電機\n(空調)", ["空調值班"])]

ws.append(["資通處通訊錄（範例，全虛構）"] + [""] * 7)
ws.append(["姓名", "", "連絡電話", "分機", "姓名", "", "連絡電話", "分機"])

rows_l, rows_r = [], []
for dept, names in STAFF_L:
    if dept != "處本部":
        rows_l.append((dept, "", ""))
    rows_l += [(n, phone(), ext()) for n in names]
for dept, names in STAFF_R:
    if not names:
        continue
    rows_r.append((dept, "", ""))
    rows_r += [(n, phone(), ext()) for n in names]

for i in range(max(len(rows_l), len(rows_r))):
    l = rows_l[i] if i < len(rows_l) else ("", "", "")
    r = rows_r[i] if i < len(rows_r) else ("", "", "")
    ws.append([l[0], "", l[1], l[2], r[0], "", r[1], r[2]])

ws.append(["廠商"] + [""] * 7)
ws.append(["公司", "", "連絡電話", "分機", "公司", "姓名", "連絡電話", "分機"])

flat_l = [(c if j == 0 else "", n, phone(), ext() if random.random() > .4 else "")
          for c, names in VENDORS_L for j, n in enumerate(names)]
flat_r = [(c if j == 0 else "", n, phone(), ext() if random.random() > .4 else "")
          for c, names in VENDORS_R for j, n in enumerate(names)]
for i in range(max(len(flat_l), len(flat_r))):
    l = flat_l[i] if i < len(flat_l) else ("", "", "", "")
    r = flat_r[i] if i < len(flat_r) else ("", "", "", "")
    ws.append([l[0], l[1], l[2], l[3], r[0], r[1], r[2], r[3]])

wb.save(os.path.join(OUT, "資通處通訊錄(範例).xlsx"))
print("通訊錄(範例).xlsx done")

# ---------- 2. 月工作日誌 ----------

HDR = ["值班日期", "值班時間", "值班人員", "事件類別", "事件描述 (Descrition)", "處理過程",
       "異常級別", "事件發生時間\nEvent Date Start", "事件處理完成時間\nEvent Date End", "處理結果"]
SHIFTS = ["夜班23-07", "早班07-15", "中班15-23"]
OPS = ["范植偉", "郭哲安", "杜冠霖", "邱繼賢"]
PATROLS = ["維護系統巡檢", "資訊機房巡檢", "資訊伺服器巡檢", "網路設備巡檢"]

INCIDENTS = {
    "202601": [
        ("服務台來電告知，一樓大廳多媒體看板黑屏。",
         "第二級\n0815 服務台來電通報多媒體看板黑屏。\n0820 通知星辰科技錢柏均，回覆派員處理。\n0950 廠商完成面板更換，畫面恢復正常。", "2"),
        ("官網無法開啟，使用者反應頁面載入逾時。",
         "第一級\n1402 服務台通報官網開啟異常。\n1405 電聯雲嶺電信洪建文確認，CDN快取異常。\n1430 廠商清除快取後恢復正常。", "1"),
        ("收到資安告警信，疑似異常登入嘗試。",
         "第一級\n2210 監控中心告警通知異常登入。\n2215 通知資安承辦確認為誤判（員工出差海外登入）。\n2240 結案。", "1"),
    ],
    "202602": [
        ("大樓側門門禁讀卡異常，人員無法刷卡進入。",
         "第二級\n0630 警衛室通報門禁讀卡異常。\n0640 電聯鼎峰工程呂偉誠，回覆讀卡機故障需更換。\n1100 更換完成測試正常。", "2"),
        ("機房空調溫度告警，A區溫度偏高。",
         "第一級\n0310 機房環控告警溫度28度。\n0315 通知青嶼電機空調值班，遠端調整後恢復。\n0400 溫度回穩23度結案。", "1"),
        ("監視器影像有雜訊，一樓大廳鏡頭畫面異常。",
         "第一級\n1520 監控中心發現監視器畫面雜訊。\n1525 通知嵐光科技江雨潔，判斷線材老化，隔日到場更換。\n次日1000 更換完成。", "1"),
    ],
}

wb2 = Workbook()
wb2.remove(wb2.active)
for month, incidents in INCIDENTS.items():
    ws2 = wb2.create_sheet(month)
    ws2.append(HDR)
    ws2.append(["範例XXXX/XX/XX", "夜班23-07", "王小明", "1.異常通報",
                "狀況敘述要點\n異常狀況：第一~三級+(說明)", "異常狀況：(時間 誰通知 異常狀況)後續處理",
                "1~3級", "2026-01-01 06:00:00", "2026-01-01 06:05:00", "1.完成"])
    y, m = int(month[:4]), int(month[4:])
    day = 1
    for inc_desc, inc_proc, level in incidents:
        for _ in range(3):  # 每筆異常前後穿插日常巡檢
            p = random.choice(PATROLS)
            t = f"{y}-{m:02d}-{day:02d}"
            ws2.append([f"{t} 00:00:00", random.choice(SHIFTS), random.choice(OPS), "6.日常巡檢", p,
                        f"記錄請參閱{p}表", "", f"{t} 07:10:00", f"{t} 07:30:00", "1.完成"])
        t = f"{y}-{m:02d}-{day:02d}"
        ws2.append([f"{t} 00:00:00", random.choice(SHIFTS), random.choice(OPS), "1.異常通報",
                    inc_desc, inc_proc, level, f"{t} 08:00:00", f"{t} 10:00:00", "1.完成"])
        day += random.randint(3, 7)

wb2.save(os.path.join(OUT, "資訊機房月工作日誌(範例).xlsx"))
print("月工作日誌(範例).xlsx done")
